import os
import openpyxl

#Excel Reader
def getRow(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row

def getCols(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row

def getCellValue(path,sheetName,rows,colums):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rows,column=colums).value

def setCellValue(path,sheetName,rows,colums,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rows, column=colums).value = data


def test_print():
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))   #used to get the absolute path
    path = os.path.join(BASE_DIR, "..", "excel", "TestData.xlsx")   #append the absolut path with file path
    work_book = openpyxl.load_workbook(path)
    sheetName = "Sheet1"

    row = getRow(path,sheetName)
    cols = getCols(path,sheetName)
    print(row,cols)

    print(getCellValue(path,sheetName,2,2))

    setCellValue(path,sheetName,2,3,"DOB")
