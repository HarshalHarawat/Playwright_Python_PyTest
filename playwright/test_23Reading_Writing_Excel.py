import os

import openpyxl
from openpyxl.reader import workbook
from playwright.sync_api import Page


def test_ReadExcelSheet(page:Page):

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))   #used to get the absolute path

    file_path = os.path.join(BASE_DIR, "..", "excel", "TestData.xlsx")   #append the absolut path with file path
    work_book = openpyxl.load_workbook(file_path)
    sheet = work_book["Sheet1"]

    row = sheet.max_row
    cols = sheet.max_column

    print(f"total number of row {row} and total number of columns {cols}")

    for i in range(1,row+1):
        for j in range(1,cols+1):
            print(sheet.cell(row=i,column=j).value)

def test_WriteExcelSheet():

    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # used to get the absolute path

    file_path = os.path.join(BASE_DIR, "..", "excel", "TestData.xlsx")  # append the absolut path with file path
    work_book = openpyxl.load_workbook(file_path)
    sheet = work_book["Sheet1"]

    for rows in range(4,8):
        for cols in range(1,4):
            sheet.cell(row=rows,column=cols).value="Hello"

    workbook.save("Path of workbook")
