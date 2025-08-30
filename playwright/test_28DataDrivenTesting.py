import openpyxl
import pytest
from playwright.sync_api import Page


def readWorkBook():

    workbook = openpyxl.load_workbook
    sheet = workbook["LoadTesting"]
    total_rows= sheet.max_row
    total_cols = sheet.max_column
    readList = []

    for row in range(2,total_rows+1):
        datalist =  []
        for cols in range(1,total_cols+1):
            data = sheet.cell(row=row,column=cols).value
            datalist.insert(cols,data)
        readList.insert(row,datalist)
    return readList


#This is how we can get the data from Excel sheet above we are taking the username and password from Excel sheet using that in below
#method

@pytest.mark.parametrize("username,password",readWorkBook())
def test_parametrization_url(page:Page,username,password):

    page.goto("https://rahulshettyacademy.com/loginpagePractise/")

    page.locator("#username").fill(username)
    page.locator("#password").fill(username)